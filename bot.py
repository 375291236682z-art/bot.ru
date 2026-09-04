import os
import json
import base64
import logging
from datetime import datetime, timedelta

import telebot
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from anthropic import Anthropic

logging.basicConfig(level=logging.INFO)
log = logging.getLogger("expense-bot")

# ---------- Настройки (из переменных окружения хостинга) ----------
BOT_TOKEN = os.environ["BOT_TOKEN"]
ANTHROPIC_API_KEY = os.environ["ANTHROPIC_API_KEY"]
# Папка для хранения файла. На Railway подключи Volume и укажи /data
DATA_DIR = os.environ.get("DATA_DIR", "/data")
# Необязательно: свой Telegram ID, чтобы ботом не пользовались чужие
OWNER_ID = os.environ.get("OWNER_ID", "").strip()

EXCEL_PATH = os.path.join(DATA_DIR, "expenses.xlsx")

CATEGORIES = [
    "Продукты", "Кафе и рестораны", "Транспорт", "Дом и быт",
    "Здоровье", "Одежда", "Развлечения", "Связь и интернет",
    "Работа", "Другое",
]

HEADERS = ["Дата", "Магазин", "Сумма", "Валюта", "Категория", "Позиции", "Добавлено"]
WIDTHS = [12, 26, 10, 9, 20, 50, 18]

bot = telebot.TeleBot(BOT_TOKEN)
claude = Anthropic(api_key=ANTHROPIC_API_KEY)


# ---------- Работа с Excel ----------
def init_excel():
    """Создаёт файл с оформленной шапкой, если его ещё нет."""
    os.makedirs(DATA_DIR, exist_ok=True)
    if os.path.exists(EXCEL_PATH):
        return
    wb = Workbook()
    ws = wb.active
    ws.title = "Расходы"
    ws.append(HEADERS)

    head_fill = PatternFill("solid", fgColor="2F5496")
    for i, _ in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=i)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = head_fill
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[get_column_letter(i)].width = WIDTHS[i - 1]
    ws.freeze_panes = "A2"
    wb.save(EXCEL_PATH)
    log.info("Создан новый файл %s", EXCEL_PATH)


def save_row(data: dict) -> None:
    init_excel()
    wb = load_workbook(EXCEL_PATH)
    ws = wb.active

    date = data.get("date") or ""
    if not date or date == "today":
        date = datetime.now().strftime("%Y-%m-%d")

    ws.append([
        date,
        data.get("shop", ""),
        float(data["total"]),
        data.get("currency", ""),
        data.get("category", "Другое"),
        (data.get("items") or "")[:200],
        datetime.now().strftime("%Y-%m-%d %H:%M"),
    ])
    ws.cell(row=ws.max_row, column=3).number_format = "#,##0.00"
    wb.save(EXCEL_PATH)


def read_rows() -> list:
    if not os.path.exists(EXCEL_PATH):
        return []
    wb = load_workbook(EXCEL_PATH, read_only=True)
    ws = wb.active
    return list(ws.iter_rows(min_row=2, values_only=True))


def is_allowed(message) -> bool:
    if not OWNER_ID:
        return True
    return str(message.from_user.id) == OWNER_ID


# ---------- Распознавание ----------
PROMPT = f"""Ты извлекаешь данные из фотографии чека.

Верни ТОЛЬКО JSON, без пояснений и без markdown-разметки, в формате:
{{
  "shop": "название магазина или пусто",
  "date": "YYYY-MM-DD (дата с чека; если не видно — today)",
  "total": число (итоговая сумма, точка как разделитель),
  "currency": "код валюты, например BYN, RUB, EUR, USD",
  "category": "одна из: {', '.join(CATEGORIES)}",
  "items": "позиции через запятую, кратко, максимум 200 символов"
}}

Если сумму разобрать невозможно, верни "total": null.
Категорию выбирай по составу покупки и типу магазина."""


def _extract_json(resp) -> dict:
    text = "".join(b.text for b in resp.content if b.type == "text")
    return json.loads(text.replace("```json", "").replace("```", "").strip())


def parse_receipt(image_bytes: bytes, mime: str = "image/jpeg") -> dict:
    resp = claude.messages.create(
        model="claude-haiku-4-5-20251001",
        max_tokens=1000,
        messages=[{
            "role": "user",
            "content": [
                {
                    "type": "image",
                    "source": {
                        "type": "base64",
                        "media_type": mime,
                        "data": base64.b64encode(image_bytes).decode(),
                    },
                },
                {"type": "text", "text": PROMPT},
            ],
        }],
    )
    return _extract_json(resp)


TEXT_PROMPT = f"""Пользователь записывает трату обычным текстом.
Верни ТОЛЬКО JSON без пояснений:
{{"shop": "описание траты", "total": число, "currency": "код валюты или пусто",
  "category": "одна из: {', '.join(CATEGORIES)}"}}
Если суммы в тексте нет, верни "total": null."""


def parse_text_expense(text: str) -> dict:
    resp = claude.messages.create(
        model="claude-haiku-4-5-20251001",
        max_tokens=500,
        messages=[{"role": "user", "content": f"{TEXT_PROMPT}\n\nТекст: {text}"}],
    )
    return _extract_json(resp)


# ---------- Команды ----------
@bot.message_handler(commands=["start", "help"])
def cmd_start(m):
    if not is_allowed(m):
        return
    bot.reply_to(m, (
        "Привет! Я веду учёт твоих расходов в Excel.\n\n"
        "📷 Пришли фото чека — распознаю и запишу\n"
        "✍️ Или напиши текстом: «кофе 8 рублей»\n\n"
        "/excel — прислать файл с расходами\n"
        "/report — расходы за последние 30 дней\n"
        "/month — расходы за текущий месяц\n"
        "/undo — удалить последнюю запись\n"
        "/id — узнать свой Telegram ID"
    ))


@bot.message_handler(commands=["id"])
def cmd_id(m):
    bot.reply_to(m, f"Твой Telegram ID: {m.from_user.id}")


@bot.message_handler(commands=["excel", "file"])
def cmd_excel(m):
    if not is_allowed(m):
        return
    if not os.path.exists(EXCEL_PATH) or not read_rows():
        bot.reply_to(m, "Пока нет ни одной записи.")
        return
    with open(EXCEL_PATH, "rb") as f:
        bot.send_document(
            m.chat.id, f,
            visible_file_name=f"Расходы_{datetime.now():%Y-%m-%d}.xlsx",
            caption="Твои расходы 📗")


@bot.message_handler(commands=["undo"])
def cmd_undo(m):
    if not is_allowed(m):
        return
    if not read_rows():
        bot.reply_to(m, "Удалять нечего.")
        return
    wb = load_workbook(EXCEL_PATH)
    ws = wb.active
    last = [ws.cell(row=ws.max_row, column=c).value for c in (2, 3, 4)]
    ws.delete_rows(ws.max_row)
    wb.save(EXCEL_PATH)
    bot.reply_to(m, f"Удалил запись: {last[0]} — {last[1]} {last[2] or ''}")


def build_report(days: int = None, this_month: bool = False) -> str:
    rows = read_rows()
    if not rows:
        return "Пока нет ни одной записи."

    if this_month:
        start = datetime.now().replace(day=1).date()
        title = "За текущий месяц"
    else:
        start = (datetime.now() - timedelta(days=days)).date()
        title = f"За последние {days} дней"

    totals, grand, currency = {}, 0.0, ""
    for r in rows:
        if not r or not r[0]:
            continue
        try:
            raw = r[0]
            d = raw.date() if isinstance(raw, datetime) else \
                datetime.strptime(str(raw)[:10], "%Y-%m-%d").date()
        except ValueError:
            continue
        if d < start:
            continue
        try:
            amount = float(str(r[2]).replace(",", "."))
        except (ValueError, TypeError):
            continue
        cat = r[4] or "Другое"
        totals[cat] = totals.get(cat, 0) + amount
        grand += amount
        if r[3]:
            currency = r[3]

    if not totals:
        return f"{title}: записей нет."

    lines = [f"📊 {title}\n"]
    for cat, s in sorted(totals.items(), key=lambda x: -x[1]):
        share = s / grand * 100 if grand else 0
        lines.append(f"{cat}: {s:.2f} {currency} ({share:.0f}%)")
    lines.append(f"\nВсего: {grand:.2f} {currency}")
    return "\n".join(lines)


@bot.message_handler(commands=["report"])
def cmd_report(m):
    if not is_allowed(m):
        return
    bot.reply_to(m, build_report(days=30))


@bot.message_handler(commands=["month"])
def cmd_month(m):
    if not is_allowed(m):
        return
    bot.reply_to(m, build_report(this_month=True))


# ---------- Фото чека ----------
@bot.message_handler(content_types=["photo"])
def handle_photo(m):
    if not is_allowed(m):
        return
    wait = bot.reply_to(m, "Читаю чек…")
    try:
        info = bot.get_file(m.photo[-1].file_id)
        image = bot.download_file(info.file_path)

        data = parse_receipt(image)
        if not data.get("total"):
            bot.edit_message_text(
                "Не смог разобрать сумму. Попробуй переснять почётче "
                "или напиши текстом, например: «продукты 42.50»",
                wait.chat.id, wait.message_id)
            return

        save_row(data)
        bot.edit_message_text(
            f"✅ Записал\n\n"
            f"🏪 {data.get('shop') or '—'}\n"
            f"💰 {data['total']} {data.get('currency', '')}\n"
            f"🏷 {data.get('category', 'Другое')}\n"
            f"📅 {data.get('date', '')}",
            wait.chat.id, wait.message_id)
    except Exception as e:
        log.exception("Ошибка обработки фото")
        bot.edit_message_text(f"Ошибка: {e}", wait.chat.id, wait.message_id)


# ---------- Текстовая трата ----------
@bot.message_handler(func=lambda m: True, content_types=["text"])
def handle_text(m):
    if not is_allowed(m):
        return
    try:
        data = parse_text_expense(m.text)
        if not data.get("total"):
            bot.reply_to(m, "Не увидел сумму. Напиши, например: «такси 15»")
            return
        save_row(data)
        bot.reply_to(
            m,
            f"✅ Записал: {data.get('shop', '')} — "
            f"{data['total']} {data.get('currency', '')} "
            f"({data.get('category', 'Другое')})")
    except Exception as e:
        log.exception("Ошибка обработки текста")
        bot.reply_to(m, f"Ошибка: {e}")


if __name__ == "__main__":
    init_excel()
    log.info("Бот запущен, файл: %s", EXCEL_PATH)
    bot.infinity_polling(skip_pending=True)
